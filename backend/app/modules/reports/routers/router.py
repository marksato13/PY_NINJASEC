import io
from datetime import datetime

from app.core.time_utils import utcnow

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.common.deps.auth import CurrentUser, require_roles
from app.common.deps.database import get_db
from app.db.models.client import Client
from app.db.models.client_profile import ClientProfile
from app.db.models.audit_log import AuditLog
from app.db.models.device import Device
from app.db.models.enums import RoleCode, TicketStatus
from app.db.models.integration import Integration
from app.db.models.report import Report
from app.db.models.report_run import ReportRun
from app.db.models.security_review import SecurityReview
from app.db.models.support_ticket import SupportTicket
from app.db.repositories.base import Repository
from app.modules.audit.services.audit_service import AuditService
from app.modules.reports.schemas import ReportCreate, ReportRead, ReportRunRead
from app.modules.reports.services.pdf_service import generate_consolidated_pdf
from app.modules.reports.services.report_service import ReportService

router = APIRouter(prefix="/reports", tags=["reports"])

_LAST_EXPORT_ACTIONS: dict[str, tuple[str, ...]] = {
    "consolidated": ("report.consolidated_export", "report.consolidated_pdf_export"),
    "devices": ("device.export",),
    "tickets": ("support_ticket.export",),
    "reviews": ("security_review.export",),
    "review-pdf": ("security_review.exported_pdf",),
    "pipeline": ("lead.export",),
}


# ---------------------------------------------------------------------------
# Rutas estáticas primero
# ---------------------------------------------------------------------------

@router.get("/consolidated-xlsx")
def export_consolidated_xlsx(
    client_id: int,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    # Verificar que el cliente pertenece a la org del usuario
    client = db.get(Client, client_id)
    if not client or client.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Client not found")

    # --- Cargar datos ---
    rev_stmt = (
        select(SecurityReview)
        .where(
            SecurityReview.organization_id == current_user.organization_id,
            SecurityReview.client_id == client_id,
        )
        .options(selectinload(SecurityReview.findings))
        .order_by(SecurityReview.scheduled_at.desc())
    )
    if date_from:
        rev_stmt = rev_stmt.where(SecurityReview.scheduled_at >= date_from)
    if date_to:
        rev_stmt = rev_stmt.where(SecurityReview.scheduled_at <= date_to)
    reviews = db.scalars(rev_stmt).all()

    tkt_stmt = select(SupportTicket).where(
        SupportTicket.organization_id == current_user.organization_id,
        SupportTicket.client_id == client_id,
    ).order_by(SupportTicket.opened_at.desc())
    if date_from:
        tkt_stmt = tkt_stmt.where(SupportTicket.opened_at >= date_from)
    if date_to:
        tkt_stmt = tkt_stmt.where(SupportTicket.opened_at <= date_to)
    tickets = db.scalars(tkt_stmt).all()

    dev_stmt = (
        select(Device)
        .join(Integration, Device.integration_id == Integration.id)
        .where(
            Integration.organization_id == current_user.organization_id,
            Integration.client_id == client_id,
        )
        .options(selectinload(Device.integration), selectinload(Device.site))
    )
    devices = db.scalars(dev_stmt).all()

    # --- Construir workbook ---
    period_label = _period_label(date_from, date_to)
    wb = _build_consolidated_workbook(client, period_label, reviews, tickets, devices)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"informe_consolidado_{client.company_name.replace(' ', '_')}_{utcnow().strftime('%Y%m%d')}.xlsx"
    AuditService(db).record(
        current_user.organization_id,
        "report.consolidated_export",
        "reports",
        f"client:{client_id}",
        current_user.id,
    )
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/consolidated-pdf")
def export_consolidated_pdf(
    client_id: int,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    client = db.get(Client, client_id)
    if not client or client.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Client not found")

    rev_stmt = (
        select(SecurityReview)
        .where(
            SecurityReview.organization_id == current_user.organization_id,
            SecurityReview.client_id == client_id,
        )
        .options(selectinload(SecurityReview.findings))
        .order_by(SecurityReview.scheduled_at.desc())
    )
    if date_from:
        rev_stmt = rev_stmt.where(SecurityReview.scheduled_at >= date_from)
    if date_to:
        rev_stmt = rev_stmt.where(SecurityReview.scheduled_at <= date_to)
    reviews = db.scalars(rev_stmt).all()

    tkt_stmt = select(SupportTicket).where(
        SupportTicket.organization_id == current_user.organization_id,
        SupportTicket.client_id == client_id,
    ).order_by(SupportTicket.opened_at.desc())
    if date_from:
        tkt_stmt = tkt_stmt.where(SupportTicket.opened_at >= date_from)
    if date_to:
        tkt_stmt = tkt_stmt.where(SupportTicket.opened_at <= date_to)
    tickets = db.scalars(tkt_stmt).all()

    dev_stmt = (
        select(Device)
        .join(Integration, Device.integration_id == Integration.id)
        .where(
            Integration.organization_id == current_user.organization_id,
            Integration.client_id == client_id,
        )
        .options(selectinload(Device.integration), selectinload(Device.site))
    )
    devices = db.scalars(dev_stmt).all()

    period_label = _period_label(date_from, date_to)
    pdf_bytes = generate_consolidated_pdf(client, period_label, reviews, tickets, devices)

    filename = f"informe_consolidado_{client.company_name.replace(' ', '_')}_{utcnow().strftime('%Y%m%d')}.pdf"
    AuditService(db).record(
        current_user.organization_id,
        "report.consolidated_pdf_export",
        "reports",
        f"client:{client_id}",
        current_user.id,
    )
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/runs", response_model=list[ReportRunRead])
def list_report_runs(
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> list[ReportRunRead]:
    if current_user.role == RoleCode.CLIENT.value:
        profile = db.scalar(
            select(ClientProfile).where(ClientProfile.user_id == current_user.id)
        )
        if not profile:
            return []
        stmt = (
            select(ReportRun)
            .join(Report, ReportRun.report_id == Report.id)
            .where(Report.client_id == profile.client_id)
        )
    else:
        stmt = (
            select(ReportRun)
            .join(Report, ReportRun.report_id == Report.id)
            .where(Report.organization_id == current_user.organization_id)
        )
    return [ReportRunRead.model_validate(item) for item in db.scalars(stmt).all()]


@router.get("/last-exports")
def get_last_exports(
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> dict[str, str | None]:
    stmt = select(AuditLog).where(
        AuditLog.organization_id == current_user.organization_id,
        AuditLog.action.in_([a for actions in _LAST_EXPORT_ACTIONS.values() for a in actions]),
    )
    logs = db.scalars(stmt).all()

    by_type: dict[str, datetime | None] = {key: None for key in _LAST_EXPORT_ACTIONS}
    for log in logs:
        for export_type, actions in _LAST_EXPORT_ACTIONS.items():
            if log.action in actions:
                current = by_type[export_type]
                if current is None or log.created_at > current:
                    by_type[export_type] = log.created_at

    return {
        key: (value.isoformat() if value is not None else None)
        for key, value in by_type.items()
    }


@router.get("/", response_model=list[ReportRead])
def list_reports(
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> list[ReportRead]:
    if current_user.role == RoleCode.CLIENT.value:
        profile = db.scalar(
            select(ClientProfile).where(ClientProfile.user_id == current_user.id)
        )
        if not profile:
            return []
        stmt = select(Report).where(Report.client_id == profile.client_id)
    else:
        stmt = select(Report).where(
            Report.organization_id == current_user.organization_id
        )
    return [ReportRead.model_validate(item) for item in db.scalars(stmt).all()]


@router.post("/", response_model=ReportRead, status_code=status.HTTP_201_CREATED)
def create_report(
    payload: ReportCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> ReportRead:
    if payload.organization_id != current_user.organization_id:
        raise HTTPException(status_code=403, detail="Organization access denied")
    report = Report(**{**payload.model_dump(), "created_by": current_user.id})
    db.add(report)
    db.commit()
    db.refresh(report)
    return ReportRead.model_validate(report)


# ---------------------------------------------------------------------------
# Rutas con path param — al final
# ---------------------------------------------------------------------------

@router.post("/{report_id}/run", response_model=ReportRunRead)
def run_report(
    report_id: int,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> ReportRunRead:
    report = Repository(db).get_by_id(Report, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if report.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Report not found")
    run = ReportService(db).generate_report_run(report)
    return ReportRunRead.model_validate(run)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _period_label(date_from: datetime | None, date_to: datetime | None) -> str:
    fmt = "%Y-%m-%d"
    if date_from and date_to:
        return f"{date_from.strftime(fmt)} — {date_to.strftime(fmt)}"
    if date_from:
        return f"Desde {date_from.strftime(fmt)}"
    if date_to:
        return f"Hasta {date_to.strftime(fmt)}"
    return "Período completo"


def _build_consolidated_workbook(
    client: Client,
    period_label: str,
    reviews: list,
    tickets: list,
    devices: list,
) -> openpyxl.Workbook:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Paleta corporativa
    COLOR_HEADER_BG  = "1E293B"   # slate-900
    COLOR_HEADER_FG  = "F8FAFC"   # slate-50
    COLOR_ALT_ROW    = "F1F5F9"   # slate-100
    COLOR_BORDER     = "CBD5E1"   # slate-300
    COLOR_ACCENT     = "0F172A"   # slate-950 para títulos de sección
    COLOR_ACCENT_FG  = "94A3B8"   # slate-400

    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    alt_fill    = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    accent_fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    header_font = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
    accent_font = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=12)
    data_font   = Font(name="Calibri", size=10)
    label_font  = Font(bold=True, name="Calibri", size=10, color=COLOR_ACCENT)

    def _header_row(ws, columns: list[tuple[str, int]]) -> None:
        for col_num, (name, width) in enumerate(columns, start=1):
            c = ws.cell(row=ws.max_row + 1 if ws.max_row else 1, column=col_num, value=name)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = width
        ws.row_dimensions[ws.max_row].height = 22

    def _data_row(ws, values: list, wrap_cols: set[int] | None = None) -> None:
        row_num = ws.max_row + 1
        fill = alt_fill if row_num % 2 == 0 else None
        for col_num, value in enumerate(values, start=1):
            c = ws.cell(row=row_num, column=col_num, value=value)
            c.font = data_font
            c.border = border
            c.alignment = Alignment(
                vertical="center",
                wrap_text=bool(wrap_cols and col_num in wrap_cols),
            )
            if fill:
                c.fill = fill

    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Hoja 1 — Resumen ejecutivo
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40

    def _kv(label: str, value) -> None:
        row = ws1.max_row + 1
        lc = ws1.cell(row=row, column=1, value=label)
        lc.font = label_font
        lc.border = border
        lc.alignment = Alignment(vertical="center")
        vc = ws1.cell(row=row, column=2, value=str(value) if value is not None else "")
        vc.font = data_font
        vc.border = border
        vc.alignment = Alignment(vertical="center", wrap_text=True)

    def _section_title(title: str) -> None:
        ws1.append([""])
        row = ws1.max_row + 1
        tc = ws1.cell(row=row, column=1, value=title)
        tc.font = accent_font
        tc.fill = accent_fill
        tc.alignment = Alignment(vertical="center")
        ws1.merge_cells(f"A{row}:B{row}")
        ws1.row_dimensions[row].height = 20

    # Encabezado principal
    title_cell = ws1.cell(row=1, column=1, value="INFORME EJECUTIVO CONSOLIDADO — NinjaSec")
    title_cell.font = Font(bold=True, name="Calibri", size=14, color=COLOR_HEADER_FG)
    title_cell.fill = accent_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells("A1:B1")
    ws1.row_dimensions[1].height = 30

    _section_title("DATOS DEL CLIENTE")
    _kv("Empresa",   client.company_name)
    _kv("Sector",    client.sector or "—")
    _kv("Tamaño",    client.size or "—")
    _kv("Estado",    str(client.commercial_status))
    _kv("Período",   period_label)
    _kv("Generado",  utcnow().strftime("%Y-%m-%d %H:%M UTC"))

    # KPIs calculados
    total_rev   = len(reviews)
    exec_rev    = sum(1 for r in reviews if r.executed_at)
    exec_pct    = round(exec_rev / total_rev * 100, 1) if total_rev else 0.0

    all_findings = [f for r in reviews for f in r.findings]
    crit_open    = sum(
        1 for f in all_findings
        if str(f.severity) in ("critical", "high") and str(f.status) == "open"
    )

    open_tkts = sum(
        1 for t in tickets
        if str(t.status) in (TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.PENDING)
    )

    _section_title("INDICADORES CLAVE")
    _kv("Revisiones programadas",          total_rev)
    _kv("Revisiones ejecutadas",           exec_rev)
    _kv("% ejecución",                     f"{exec_pct}%")
    _kv("Hallazgos totales",               len(all_findings))
    _kv("Hallazgos críticos/altos abiertos", crit_open)
    _kv("Tickets totales",                 len(tickets))
    _kv("Tickets abiertos",                open_tkts)
    _kv("Activos registrados",             len(devices))

    ws1.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Hoja 2 — Revisiones
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Revisiones")
    rev_cols = [
        ("ID",                8),
        ("Estado",           14),
        ("Revisor ID",       12),
        ("Fecha programada", 22),
        ("Fecha ejecución",  22),
        ("Hallazgos",         10),
        ("Notas",            40),
    ]
    _header_row(ws2, rev_cols)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(rev_cols))}1"

    for r in reviews:
        _data_row(ws2, [
            r.id,
            str(r.status),
            r.reviewer_user_id or "",
            r.scheduled_at.strftime("%Y-%m-%d %H:%M") if r.scheduled_at else "",
            r.executed_at.strftime("%Y-%m-%d %H:%M") if r.executed_at else "",
            len(r.findings),
            r.notes or "",
        ], wrap_cols={7})

    # -----------------------------------------------------------------------
    # Hoja 3 — Hallazgos
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Hallazgos")
    fnd_cols = [
        ("ID",           7),
        ("Revisión ID", 12),
        ("Severidad",   14),
        ("Título",      35),
        ("Estado",      14),
        ("Descripción", 45),
    ]
    _header_row(ws3, fnd_cols)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(fnd_cols))}1"

    for f in all_findings:
        _data_row(ws3, [
            f.id,
            f.review_id,
            str(f.severity),
            f.title,
            str(f.status),
            f.description or "",
        ], wrap_cols={6})

    # -----------------------------------------------------------------------
    # Hoja 4 — Tickets
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Tickets")
    tkt_cols = [
        ("ID",             7),
        ("Título",        35),
        ("Prioridad",     12),
        ("Estado",        14),
        ("Días abierto",  12),
        ("Asignado a",    12),
        ("Fecha apertura",20),
        ("Fecha cierre",  20),
        ("Resolución",    45),
    ]
    _header_row(ws4, tkt_cols)
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(tkt_cols))}1"

    now = utcnow()
    for t in tickets:
        ref_date = t.closed_at or now
        days_open = (ref_date - t.opened_at).days if t.opened_at else ""
        _data_row(ws4, [
            t.id,
            t.title,
            str(t.priority),
            str(t.status),
            days_open,
            t.assigned_to or "",
            t.opened_at.strftime("%Y-%m-%d %H:%M") if t.opened_at else "",
            t.closed_at.strftime("%Y-%m-%d %H:%M") if t.closed_at else "",
            t.resolution or "",
        ], wrap_cols={9})

    # -----------------------------------------------------------------------
    # Hoja 5 — Inventario
    # -----------------------------------------------------------------------
    ws5 = wb.create_sheet("Inventario")
    inv_cols = [
        ("ID",          7),
        ("Hostname",   22),
        ("Tipo",       14),
        ("Vendor",     14),
        ("IP",         16),
        ("S/N",        16),
        ("Asset Tag",  14),
        ("Sede",       18),
        ("Consola",    22),
        ("Estado",     12),
    ]
    _header_row(ws5, inv_cols)
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = f"A1:{get_column_letter(len(inv_cols))}1"

    for d in devices:
        _data_row(ws5, [
            d.id,
            d.hostname,
            d.device_type or "",
            d.vendor or "",
            d.ip_address or "",
            d.serial_number or "",
            d.asset_tag or "",
            d.site.name if d.site else "",
            d.integration.name if d.integration else "",
            str(d.status),
        ])

    return wb
