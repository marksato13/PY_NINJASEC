import io
from datetime import datetime

from app.core.time_utils import utcnow

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.common.deps.auth import CurrentUser, require_roles
from app.common.deps.database import get_db
from app.db.models.enums import RoleCode, TicketPriority, TicketStatus
from app.db.models.support_ticket import SupportTicket
from app.modules.audit.services.audit_service import AuditService
from app.modules.support_tickets.schemas import (
    SupportTicketCreate,
    SupportTicketDetail,
    SupportTicketRead,
    SupportTicketUpdate,
    TicketEventCreate,
    TicketEventRead,
    TicketFromFindingCreate,
    TicketImportResult,
    TicketStats,
)
from app.modules.support_tickets.services.ticket_service import TicketService

router = APIRouter(prefix="/support-tickets", tags=["support-tickets"])

_VALID_PRIORITIES = {v.value for v in TicketPriority}
_IMPORT_COLS = ["client_id", "title", "priority", "description", "category", "integration_id", "device_id"]


# ---------------------------------------------------------------------------
# Rutas estáticas primero (deben ir ANTES de /{ticket_id})
# ---------------------------------------------------------------------------

@router.get("/stats", response_model=TicketStats)
def get_ticket_stats(
    client_id: int | None = None,
    date_from: datetime | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> TicketStats:
    return TicketService(db, current_user).stats(client_id, date_from)


@router.get("/export-xlsx")
def export_tickets_xlsx(
    client_id: int | None = None,
    ticket_status: str | None = None,
    priority: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    svc = TicketService(db, current_user)
    tickets = svc.query_for_export(client_id, ticket_status, priority, date_from, date_to)

    wb = _build_tickets_workbook(tickets)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tickets_{utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    AuditService(db).record(
        current_user.organization_id,
        "support_ticket.export",
        "support_tickets",
        f"count:{len(tickets)}",
        current_user.id,
    )
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Listado general
# ---------------------------------------------------------------------------

@router.get("/", response_model=list[SupportTicketRead])
def list_tickets(
    client_id: int | None = None,
    ticket_status: str | None = None,
    priority: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> list[SupportTicketRead]:
    tickets = TicketService(db, current_user).list(client_id, ticket_status, priority, date_from, date_to)
    return [SupportTicketRead.model_validate(t) for t in tickets]


# ---------------------------------------------------------------------------
# Crear
# ---------------------------------------------------------------------------

@router.post("/", response_model=SupportTicketRead, status_code=status.HTTP_201_CREATED)
def create_ticket(
    payload: SupportTicketCreate,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> SupportTicketRead:
    if current_user.role == RoleCode.CLIENT.value:
        from sqlalchemy import select as sa_select
        from app.db.models.client_profile import ClientProfile
        profile = db.scalar(sa_select(ClientProfile).where(ClientProfile.user_id == current_user.id))
        if not profile:
            raise HTTPException(status_code=403, detail="No client profile associated with this account")
        payload = payload.model_copy(update={"client_id": profile.client_id})
    elif payload.client_id is None:
        raise HTTPException(status_code=422, detail="client_id is required")
    return SupportTicketRead.model_validate(TicketService(db, current_user).create(payload))


@router.post("/import", response_model=TicketImportResult)
async def import_tickets_xlsx(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> TicketImportResult:
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    raw_headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    col_idx: dict[str, int] = {col: raw_headers.index(col) for col in _IMPORT_COLS if col in raw_headers}

    missing_required = {"client_id", "title", "priority"} - col_idx.keys()
    if missing_required:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas requeridas faltantes: {', '.join(missing_required)}",
        )

    created = 0
    skipped = 0
    errors: list[str] = []

    def _cell(row: tuple, col: str):
        idx = col_idx.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    for row_num, row in enumerate(rows[1:], start=2):
        client_id_val = _cell(row, "client_id")
        title_val = _cell(row, "title")
        priority_val = str(_cell(row, "priority") or "").strip().lower()

        if not client_id_val or not title_val:
            errors.append(f"Fila {row_num}: client_id y title son obligatorios")
            skipped += 1
            continue

        if priority_val not in _VALID_PRIORITIES:
            errors.append(f"Fila {row_num}: priority '{priority_val}' no válido — usar: {', '.join(_VALID_PRIORITIES)}")
            skipped += 1
            continue

        try:
            db.add(SupportTicket(
                organization_id=current_user.organization_id,
                client_id=int(client_id_val),
                title=str(title_val).strip(),
                description=str(_cell(row, "description") or "").strip() or None,
                category=str(_cell(row, "category") or "").strip() or None,
                priority=TicketPriority(priority_val),
                status=TicketStatus.OPEN,
                integration_id=int(_cell(row, "integration_id")) if _cell(row, "integration_id") else None,
                device_id=int(_cell(row, "device_id")) if _cell(row, "device_id") else None,
            ))
            created += 1
        except Exception as exc:
            errors.append(f"Fila {row_num}: {exc}")
            skipped += 1

    if created:
        db.commit()
        AuditService(db).record(
            current_user.organization_id,
            "support_ticket.import",
            "support_tickets",
            f"batch:{created}",
            current_user.id,
        )

    return TicketImportResult(created=created, skipped=skipped, errors=errors)


# ---------------------------------------------------------------------------
# Crear ticket desde hallazgo — ruta estática antes de /{ticket_id}
# ---------------------------------------------------------------------------

@router.post(
    "/from-finding/{finding_id}",
    response_model=SupportTicketRead,
    status_code=status.HTTP_201_CREATED,
)
def create_ticket_from_finding(
    finding_id: int,
    payload: TicketFromFindingCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> SupportTicketRead:
    return SupportTicketRead.model_validate(
        TicketService(db, current_user).create_from_finding(finding_id, payload)
    )


# ---------------------------------------------------------------------------
# Rutas con path param /{ticket_id} — siempre al final
# ---------------------------------------------------------------------------

@router.get("/{ticket_id}", response_model=SupportTicketDetail)
def get_ticket(
    ticket_id: int,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> SupportTicketDetail:
    return SupportTicketDetail.model_validate(TicketService(db, current_user).get_or_404(ticket_id))


@router.patch("/{ticket_id}", response_model=SupportTicketRead)
def update_ticket(
    ticket_id: int,
    payload: SupportTicketUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> SupportTicketRead:
    return SupportTicketRead.model_validate(TicketService(db, current_user).update(ticket_id, payload))


@router.delete("/{ticket_id}")
def delete_ticket(
    ticket_id: int,
    current_user: CurrentUser = Depends(require_roles(RoleCode.SUPER_ADMIN.value)),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    TicketService(db, current_user).delete(ticket_id)
    return {"message": "Ticket deleted"}


@router.post("/{ticket_id}/events", response_model=TicketEventRead, status_code=status.HTTP_201_CREATED)
def add_ticket_event(
    ticket_id: int,
    payload: TicketEventCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> TicketEventRead:
    return TicketEventRead.model_validate(TicketService(db, current_user).add_event(ticket_id, payload))


# ---------------------------------------------------------------------------
# Helper: construir workbook con plantilla corporativa NinjaSec
# ---------------------------------------------------------------------------

def _build_tickets_workbook(tickets: list) -> openpyxl.Workbook:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    COLOR_HEADER_BG = "1E293B"
    COLOR_HEADER_FG = "F8FAFC"
    COLOR_ALT_ROW   = "F1F5F9"
    COLOR_BORDER    = "CBD5E1"

    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    alt_fill    = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    header_font = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
    data_font   = Font(name="Calibri", size=10)

    columns = [
        ("ID",              8),
        ("Cliente ID",     12),
        ("Título",         40),
        ("Categoría",      18),
        ("Prioridad",      12),
        ("Estado",         14),
        ("Asignado a",     14),
        ("Fecha apertura", 20),
        ("Fecha cierre",   20),
        ("Resolución",     50),
    ]

    for col_num, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_num, t in enumerate(tickets, start=2):
        fill = alt_fill if row_num % 2 == 0 else None
        values = [
            t.id, t.client_id, t.title, t.category or "",
            str(t.priority), str(t.status), t.assigned_to or "",
            t.opened_at.strftime("%Y-%m-%d %H:%M") if t.opened_at else "",
            t.closed_at.strftime("%Y-%m-%d %H:%M") if t.closed_at else "",
            t.resolution or "",
        ]
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_num == 10))
            if fill:
                cell.fill = fill

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}1"

    return wb
