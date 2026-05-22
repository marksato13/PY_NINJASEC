import io
from datetime import datetime

from app.core.time_utils import utcnow

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.common.deps.auth import CurrentUser, require_roles
from app.common.deps.database import get_db
from app.db.models.client_profile import ClientProfile
from app.db.models.enums import (
    FindingSeverity,
    FindingStatus,
    RecommendationStatus,
    ReviewStatus,
    RoleCode,
    TicketPriority,
    TicketStatus,
)
from app.db.models.security_review import (
    ReviewAttachment,
    ReviewChecklistItem,
    ReviewFinding,
    ReviewRecommendation,
    SecurityReview,
)
from app.db.models.support_ticket import SupportTicket
from app.db.repositories.base import Repository
from app.modules.audit.services.audit_service import AuditService
from app.modules.security_reviews.schemas import (
    AttachmentCreate,
    AttachmentRead,
    ChecklistItemCreate,
    ChecklistItemRead,
    FindingCreate,
    FindingRead,
    FindingUpdate,
    RecommendationCreate,
    RecommendationRead,
    SecurityReviewCreate,
    SecurityReviewDetail,
    SecurityReviewRead,
    SecurityReviewUpdate,
)

router = APIRouter(prefix="/security-reviews", tags=["security-reviews"])


@router.get("/", response_model=list[SecurityReviewRead])
def list_reviews(
    client_id: int | None = None,
    integration_id: int | None = None,
    status: str | None = None,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> list[SecurityReviewRead]:
    stmt = select(SecurityReview).where(SecurityReview.organization_id == current_user.organization_id)
    if current_user.role == RoleCode.CLIENT.value:
        profile = db.scalar(
            select(ClientProfile).where(ClientProfile.user_id == current_user.id)
        )
        if not profile:
            return []
        stmt = stmt.where(
            SecurityReview.client_id == profile.client_id,
            SecurityReview.status == ReviewStatus.CLOSED,
        )
    elif current_user.role == RoleCode.COLLABORATOR.value:
        # Collaborators solo ven revisiones donde son el revisor asignado
        stmt = stmt.where(SecurityReview.reviewer_user_id == current_user.id)
        if client_id:
            stmt = stmt.where(SecurityReview.client_id == client_id)
    elif client_id:
        stmt = stmt.where(SecurityReview.client_id == client_id)
    if integration_id:
        stmt = stmt.where(SecurityReview.integration_id == integration_id)
    if status:
        stmt = stmt.where(SecurityReview.status == status)
    reviews = list(db.scalars(stmt).all())
    if reviews:
        review_ids = [r.id for r in reviews]
        total_counts = dict(
            db.execute(
                select(ReviewFinding.review_id, func.count(ReviewFinding.id))
                .where(ReviewFinding.review_id.in_(review_ids))
                .group_by(ReviewFinding.review_id)
            ).all()
        )
        critical_counts = dict(
            db.execute(
                select(ReviewFinding.review_id, func.count(ReviewFinding.id))
                .where(ReviewFinding.review_id.in_(review_ids))
                .where(ReviewFinding.severity == FindingSeverity.CRITICAL)
                .group_by(ReviewFinding.review_id)
            ).all()
        )
        for r in reviews:
            r.findings_count = total_counts.get(r.id, 0)
            r.critical_findings_count = critical_counts.get(r.id, 0)
    return [SecurityReviewRead.model_validate(r) for r in reviews]


@router.post("/", response_model=SecurityReviewRead, status_code=status.HTTP_201_CREATED)
def create_review(
    payload: SecurityReviewCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> SecurityReviewRead:
    review = SecurityReview(
        organization_id=current_user.organization_id,
        client_id=payload.client_id,
        integration_id=payload.integration_id,
        scheduled_at=payload.scheduled_at,
        reviewer_user_id=payload.reviewer_user_id,
        notes=payload.notes,
        status=ReviewStatus.SCHEDULED,
    )
    db.add(review)
    db.commit()
    db.refresh(review)
    AuditService(db).record(
        current_user.organization_id, "security_review.created", "security_reviews", str(review.id), current_user.id
    )
    return SecurityReviewRead.model_validate(review)


# ---------------------------------------------------------------------------
# Export XLSX — ruta estática ANTES de /{review_id}  (RF-17)
# ---------------------------------------------------------------------------

@router.get("/export-xlsx")
def export_reviews_xlsx(
    client_id: int | None = None,
    review_status: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    stmt = (
        select(SecurityReview)
        .where(SecurityReview.organization_id == current_user.organization_id)
        .options(
            selectinload(SecurityReview.findings).selectinload(ReviewFinding.recommendations),
            selectinload(SecurityReview.recommendations),
        )
    )
    if client_id:
        stmt = stmt.where(SecurityReview.client_id == client_id)
    if review_status:
        stmt = stmt.where(SecurityReview.status == review_status)
    if date_from:
        stmt = stmt.where(SecurityReview.scheduled_at >= date_from)
    if date_to:
        stmt = stmt.where(SecurityReview.scheduled_at <= date_to)
    stmt = stmt.order_by(SecurityReview.scheduled_at.desc())

    reviews = db.scalars(stmt).all()

    wb = _build_reviews_workbook(reviews)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"revisiones_{utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    AuditService(db).record(
        current_user.organization_id,
        "security_review.export",
        "security_reviews",
        f"count:{len(reviews)}",
        current_user.id,
    )
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{review_id}/pdf")
def export_review_pdf(
    review_id: int,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    from app.db.models.client import Client
    from app.db.models.integration import Integration
    from app.db.models.user import User
    from app.modules.reports.services.pdf_service import generate_security_review_pdf

    stmt = (
        select(SecurityReview)
        .where(
            SecurityReview.id == review_id,
            SecurityReview.organization_id == current_user.organization_id,
        )
        .options(
            selectinload(SecurityReview.checklist_items),
            selectinload(SecurityReview.findings),
            selectinload(SecurityReview.recommendations),
        )
    )
    review = db.scalars(stmt).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    # Collaborator solo puede exportar revisiones asignadas
    if current_user.role == RoleCode.COLLABORATOR.value and review.reviewer_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Collaborators can only export their assigned reviews")
    # Cliente solo accede a revisiones cerradas de su propia empresa
    if current_user.role == RoleCode.CLIENT.value:
        profile = db.scalar(select(ClientProfile).where(ClientProfile.user_id == current_user.id))
        if not profile or profile.client_id != review.client_id or str(review.status) != "closed":
            raise HTTPException(status_code=404, detail="Review not found")

    client = db.get(Client, review.client_id)
    integration = db.get(Integration, review.integration_id)
    reviewer_name = None
    if review.reviewer_user_id:
        reviewer = db.get(User, review.reviewer_user_id)
        if reviewer:
            reviewer_name = reviewer.full_name

    pdf_bytes = generate_security_review_pdf(
        review=review,
        client=client,
        integration=integration,
        reviewer_name=reviewer_name,
        findings=list(review.findings or []),
        checklist_items=list(review.checklist_items or []),
        recommendations=list(review.recommendations or []),
    )
    AuditService(db).record(
        current_user.organization_id,
        "security_review.exported_pdf",
        "security_reviews",
        str(review.id),
        current_user.id,
    )
    filename = f"revision_{review.id}_{utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{review_id}", response_model=SecurityReviewDetail)
def get_review(
    review_id: int,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> SecurityReviewDetail:
    stmt = (
        select(SecurityReview)
        .where(SecurityReview.id == review_id, SecurityReview.organization_id == current_user.organization_id)
        .options(
            selectinload(SecurityReview.checklist_items),
            selectinload(SecurityReview.findings),
            selectinload(SecurityReview.recommendations),
            selectinload(SecurityReview.attachments),
        )
    )
    review = db.scalar(stmt)
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    if current_user.role == RoleCode.CLIENT.value:
        profile = db.scalar(
            select(ClientProfile).where(ClientProfile.user_id == current_user.id)
        )
        if (
            not profile
            or review.client_id != profile.client_id
            or review.status != ReviewStatus.CLOSED
        ):
            raise HTTPException(status_code=404, detail="Review not found")
    return SecurityReviewDetail.model_validate(review)


@router.delete("/{review_id}")
def delete_review(
    review_id: int,
    current_user: CurrentUser = Depends(require_roles(RoleCode.SUPER_ADMIN.value)),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    review = Repository(db).get_by_id(SecurityReview, review_id)
    if not review or review.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Review not found")
    db.delete(review)
    db.commit()
    AuditService(db).record(
        current_user.organization_id, "security_review.deleted", "security_reviews", str(review_id), current_user.id
    )
    return {"message": "Review deleted"}


@router.patch("/{review_id}", response_model=SecurityReviewRead)
def update_review(
    review_id: int,
    payload: SecurityReviewUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> SecurityReviewRead:
    review = Repository(db).get_by_id(SecurityReview, review_id)
    if not review or review.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Review not found")
    if current_user.role == RoleCode.COLLABORATOR.value and review.reviewer_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Collaborators can only edit reviews assigned to them")
    data = payload.model_dump(exclude_unset=True)
    # Validar: cerrar requiere reviewer asignado
    new_status = data.get("status") or review.status
    reviewer = data.get("reviewer_user_id") or review.reviewer_user_id
    if new_status == ReviewStatus.CLOSED and not reviewer:
        raise HTTPException(status_code=422, detail="reviewer_user_id is required to close a review")
    for field, value in data.items():
        if field == "status" and value is not None:
            value = ReviewStatus(value)
        setattr(review, field, value)
    db.add(review)
    db.commit()
    db.refresh(review)
    AuditService(db).record(
        current_user.organization_id, "security_review.updated", "security_reviews", str(review.id), current_user.id
    )
    return SecurityReviewRead.model_validate(review)


# --- Checklist ---

@router.post("/{review_id}/checklist", response_model=ChecklistItemRead, status_code=status.HTTP_201_CREATED)
def add_checklist_item(
    review_id: int,
    payload: ChecklistItemCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> ChecklistItemRead:
    review = Repository(db).get_by_id(SecurityReview, review_id)
    if not review or review.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Review not found")
    item = ReviewChecklistItem(review_id=review_id, **payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return ChecklistItemRead.model_validate(item)


# --- Findings ---

@router.post("/{review_id}/findings", response_model=FindingRead, status_code=status.HTTP_201_CREATED)
def add_finding(
    review_id: int,
    payload: FindingCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> FindingRead:
    review = Repository(db).get_by_id(SecurityReview, review_id)
    if not review or review.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Review not found")
    severity = FindingSeverity(payload.severity)
    requires_follow_up = severity in {
        FindingSeverity.CRITICAL,
        FindingSeverity.HIGH,
    }
    has_recommendation = bool(payload.recommendation and payload.recommendation.strip())
    has_ticket = payload.create_ticket
    if requires_follow_up and not (has_recommendation or has_ticket):
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Critical/high findings require a ticket or recommendation",
                "action_required": True,
            },
        )
    if has_ticket and not payload.ticket_priority:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "ticket_priority is required when create_ticket is true",
                "action_required": True,
            },
        )
    finding = ReviewFinding(
        review_id=review_id,
        severity=severity,
        title=payload.title,
        description=payload.description,
        evidence_url=payload.evidence_url,
        status=FindingStatus.OPEN,
    )
    db.add(finding)
    db.flush()

    if has_recommendation:
        db.add(
            ReviewRecommendation(
                review_id=review_id,
                finding_id=finding.id,
                recommendation=payload.recommendation.strip(),
                owner_user_id=payload.recommendation_owner_user_id,
                due_date=payload.recommendation_due_date,
                status=RecommendationStatus.PENDING,
            )
        )

    if has_ticket:
        db.add(
            SupportTicket(
                organization_id=current_user.organization_id,
                client_id=review.client_id,
                integration_id=review.integration_id,
                review_id=review.id,
                finding_id=finding.id,
                title=f"Hallazgo: {finding.title}",
                description=payload.ticket_description_override or finding.description,
                category=payload.ticket_category,
                priority=TicketPriority(payload.ticket_priority),
                status=TicketStatus.OPEN,
                assigned_to=payload.ticket_assigned_to,
            )
        )

    db.commit()
    db.refresh(finding)
    AuditService(db).record(
        current_user.organization_id, "review_finding.created", "review_findings", str(finding.id), current_user.id
    )
    return FindingRead.model_validate(finding)


@router.patch("/findings/{finding_id}", response_model=FindingRead)
def update_finding(
    finding_id: int,
    payload: FindingUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> FindingRead:
    finding = Repository(db).get_by_id(ReviewFinding, finding_id)
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field == "severity" and value is not None:
            value = FindingSeverity(value)
        if field == "status" and value is not None:
            value = FindingStatus(value)
        setattr(finding, field, value)
    db.add(finding)
    db.commit()
    db.refresh(finding)
    return FindingRead.model_validate(finding)


# --- Recommendations ---

@router.post("/{review_id}/recommendations", response_model=RecommendationRead, status_code=status.HTTP_201_CREATED)
def add_recommendation(
    review_id: int,
    payload: RecommendationCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> RecommendationRead:
    review = Repository(db).get_by_id(SecurityReview, review_id)
    if not review or review.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Review not found")
    rec = ReviewRecommendation(
        review_id=review_id,
        finding_id=payload.finding_id,
        recommendation=payload.recommendation,
        owner_user_id=payload.owner_user_id,
        due_date=payload.due_date,
        status=RecommendationStatus.PENDING,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return RecommendationRead.model_validate(rec)


# --- Attachments ---

@router.post("/{review_id}/attachments", response_model=AttachmentRead, status_code=status.HTTP_201_CREATED)
def add_attachment(
    review_id: int,
    payload: AttachmentCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> AttachmentRead:
    review = Repository(db).get_by_id(SecurityReview, review_id)
    if not review or review.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Review not found")
    attachment = ReviewAttachment(
        review_id=review_id,
        finding_id=payload.finding_id,
        file_url=payload.file_url,
        filename=payload.filename,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    return AttachmentRead.model_validate(attachment)


# ---------------------------------------------------------------------------
# Helper: workbook con 3 hojas — plantilla corporativa NinjaSec
# ---------------------------------------------------------------------------

def _build_reviews_workbook(reviews: list) -> openpyxl.Workbook:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    COLOR_HEADER_BG = "1E293B"
    COLOR_HEADER_FG = "F8FAFC"
    COLOR_ALT_ROW   = "F1F5F9"
    COLOR_BORDER    = "CBD5E1"

    def _make_styles():
        thin = Side(style="thin", color=COLOR_BORDER)
        return (
            Border(left=thin, right=thin, top=thin, bottom=thin),
            PatternFill("solid", fgColor=COLOR_HEADER_BG),
            PatternFill("solid", fgColor=COLOR_ALT_ROW),
            Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11),
            Font(name="Calibri", size=10),
        )

    def _write_header(ws, columns):
        border, header_fill, _, header_font, _ = _make_styles()
        for col_num, (col_name, col_width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = col_width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    def _write_row(ws, row_num, values, wrap_cols: set[int] | None = None):
        border, _, alt_fill, _, data_font = _make_styles()
        fill = alt_fill if row_num % 2 == 0 else None
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=bool(wrap_cols and col_num in wrap_cols),
            )
            if fill:
                cell.fill = fill

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Hoja 1 — Revisiones
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Revisiones"
    review_cols = [
        ("ID",                 7),
        ("Cliente ID",        12),
        ("Consola ID",        12),
        ("Estado",            14),
        ("Revisor ID",        12),
        ("Fecha programada",  22),
        ("Fecha ejecución",   22),
        ("Notas",             40),
    ]
    _write_header(ws1, review_cols)

    for row_num, r in enumerate(reviews, start=2):
        _write_row(ws1, row_num, [
            r.id,
            r.client_id,
            r.integration_id,
            str(r.status),
            r.reviewer_user_id or "",
            r.scheduled_at.strftime("%Y-%m-%d %H:%M") if r.scheduled_at else "",
            r.executed_at.strftime("%Y-%m-%d %H:%M") if r.executed_at else "",
            r.notes or "",
        ], wrap_cols={8})

    # ------------------------------------------------------------------
    # Hoja 2 — Hallazgos
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Hallazgos")
    finding_cols = [
        ("ID",            7),
        ("Revisión ID",  12),
        ("Severidad",    14),
        ("Título",       35),
        ("Estado",       14),
        ("Descripción",  45),
        ("Evidencia URL",35),
    ]
    _write_header(ws2, finding_cols)

    row_num = 2
    for r in reviews:
        for f in r.findings:
            _write_row(ws2, row_num, [
                f.id,
                f.review_id,
                str(f.severity),
                f.title,
                str(f.status),
                f.description or "",
                f.evidence_url or "",
            ], wrap_cols={6})
            row_num += 1

    # ------------------------------------------------------------------
    # Hoja 3 — Recomendaciones
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Recomendaciones")
    rec_cols = [
        ("ID",             7),
        ("Revisión ID",   12),
        ("Hallazgo ID",   12),
        ("Recomendación", 50),
        ("Responsable ID",14),
        ("Fecha límite",  20),
        ("Estado",        14),
    ]
    _write_header(ws3, rec_cols)

    row_num = 2
    for r in reviews:
        for rec in r.recommendations:
            _write_row(ws3, row_num, [
                rec.id,
                rec.review_id,
                rec.finding_id or "",
                rec.recommendation,
                rec.owner_user_id or "",
                rec.due_date.strftime("%Y-%m-%d") if rec.due_date else "",
                str(rec.status),
            ], wrap_cols={4})
            row_num += 1

    return wb
