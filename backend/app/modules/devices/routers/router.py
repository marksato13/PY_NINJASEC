import io

from app.core.time_utils import utcnow

import openpyxl
from fastapi import APIRouter, Depends, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.common.deps.auth import CurrentUser, require_roles
from app.common.deps.database import get_db
from app.db.models.enums import RoleCode
from app.modules.audit.services.audit_service import AuditService
from app.modules.devices.schemas import DeviceCreate, DeviceRead, DeviceUpdate, TopologyResponse
from app.modules.devices.services.device_service import DeviceService
from app.modules.devices.services.topology_service import TopologyService

router = APIRouter(prefix="/devices", tags=["devices"])


# ---------------------------------------------------------------------------
# Rutas estáticas primero (antes de /{device_id})
# ---------------------------------------------------------------------------

@router.get("/export-xlsx")
def export_devices_xlsx(
    client_id: int | None = None,
    integration_id: int | None = None,
    site_id: int | None = None,
    device_status: str | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    svc = DeviceService(db, current_user)
    devices = svc.query_for_export(client_id, integration_id, site_id, device_status)

    wb = _build_devices_workbook(devices)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventario_{utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    AuditService(db).record(
        current_user.organization_id,
        "device.export",
        "devices",
        f"count:{len(devices)}",
        current_user.id,
    )
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Topología por cliente
# ---------------------------------------------------------------------------

@router.get("/topology", response_model=TopologyResponse)
def get_topology(
    client_id: int,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
        )
    ),
    db: Session = Depends(get_db),
) -> TopologyResponse:
    """Devuelve nodes + edges + clusters para diagramar la topología del cliente."""
    return TopologyService(db, current_user).build(client_id)


# ---------------------------------------------------------------------------
# Listado y creación
# ---------------------------------------------------------------------------

@router.get("/", response_model=list[DeviceRead])
def list_devices(
    client_id: int | None = None,
    integration_id: int | None = None,
    site_id: int | None = None,
    device_status: str | None = None,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> list[DeviceRead]:
    devices = DeviceService(db, current_user).list(client_id, integration_id, site_id, device_status)
    return [DeviceRead.model_validate(d) for d in devices]


@router.post("/", response_model=DeviceRead, status_code=status.HTTP_201_CREATED)
def create_device(
    payload: DeviceCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> DeviceRead:
    return DeviceRead.model_validate(DeviceService(db, current_user).create(payload))


# ---------------------------------------------------------------------------
# Rutas con path param — siempre al final
# ---------------------------------------------------------------------------

@router.get("/{device_id}", response_model=DeviceRead)
def get_device(
    device_id: int,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> DeviceRead:
    return DeviceRead.model_validate(DeviceService(db, current_user).get_or_404(device_id))


@router.patch("/{device_id}", response_model=DeviceRead)
def update_device(
    device_id: int,
    payload: DeviceUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value, RoleCode.COLLABORATOR.value)
    ),
    db: Session = Depends(get_db),
) -> DeviceRead:
    return DeviceRead.model_validate(DeviceService(db, current_user).update(device_id, payload))


@router.delete("/{device_id}", response_model=DeviceRead)
def retire_device(
    device_id: int,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> DeviceRead:
    """Soft-delete — marca el activo como 'retired', no elimina físicamente."""
    return DeviceRead.model_validate(DeviceService(db, current_user).retire(device_id))


# ---------------------------------------------------------------------------
# Helper: workbook inventario con plantilla corporativa
# ---------------------------------------------------------------------------

def _build_devices_workbook(devices: list) -> openpyxl.Workbook:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    COLOR_HEADER_BG = "1E293B"
    COLOR_HEADER_FG = "F8FAFC"
    COLOR_ALT_ROW   = "F1F5F9"
    COLOR_BORDER    = "CBD5E1"

    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    alt_fill    = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    header_font = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
    data_font   = Font(name="Calibri", size=10)

    columns = [
        ("ID",            7),
        ("Hostname",     22),
        ("Tipo",         14),
        ("Vendor",       14),
        ("Modelo",       16),
        ("IP",           16),
        ("S/N",          16),
        ("Asset Tag",    14),
        ("Propietario",  20),
        ("Sede",         18),
        ("Consola",      22),
        ("Estado",       12),
        ("Última vista", 20),
    ]

    for col_num, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_num, d in enumerate(devices, start=2):
        fill = alt_fill if row_num % 2 == 0 else None
        site_name    = d.site.name if d.site else ""
        console_name = d.integration.name if d.integration else "Sin integración"
        last_seen    = d.last_seen_at.strftime("%Y-%m-%d %H:%M") if d.last_seen_at else ""

        values = [
            d.id, d.hostname, d.device_type or "", d.vendor or "", d.model or "",
            d.ip_address or "", d.serial_number or "", d.asset_tag or "",
            d.device_owner or "", site_name, console_name, str(d.status), last_seen,
        ]
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}1"

    return wb
