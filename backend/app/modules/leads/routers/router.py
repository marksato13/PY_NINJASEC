import io

from app.core.time_utils import utcnow

import openpyxl
from fastapi import APIRouter, Depends, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.common.deps.auth import CurrentUser, require_roles
from app.common.deps.database import get_db
from app.db.models.enums import RoleCode
from app.modules.leads.schemas import LeadCreate, LeadInfoUpdate, LeadRead, LeadUpdate
from app.modules.leads.services.lead_service import LeadService

router = APIRouter(prefix="/leads", tags=["leads"])


@router.get("/export-xlsx")
def export_leads_xlsx(
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    leads = LeadService(db, current_user).list()
    # Filtro por rango de fechas sobre created_at
    if date_from or date_to:
        from datetime import datetime as _dt
        df = _dt.fromisoformat(date_from) if date_from else None
        dt = _dt.fromisoformat(date_to)   if date_to   else None
        if dt:
            dt = dt.replace(hour=23, minute=59, second=59)
        leads = [
            lead for lead in leads
            if lead.created_at
            and (df is None or lead.created_at >= df)
            and (dt is None or lead.created_at <= dt)
        ]

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1E293B")
    h_font = Font(bold=True, color="F8FAFC", name="Calibri", size=11)
    d_font = Font(name="Calibri", size=10)
    a_fill = PatternFill("solid", fgColor="F1F5F9")

    columns = [
        ("ID", 6), ("Estado", 14), ("Empresa", 24), ("Contacto", 22),
        ("Email", 28), ("Teléfono", 16), ("Área de interés", 22),
        ("Fuente", 12), ("Mensaje", 40), ("Creado", 18),
    ]
    for i, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = h_font; cell.fill = h_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_num, lead in enumerate(leads, 2):
        fill = a_fill if row_num % 2 == 0 else None
        values = [
            lead.id, str(lead.status), lead.company_name or "",
            lead.contact_name, lead.email, lead.phone or "",
            lead.interest_area or "", lead.source or "",
            lead.message or "",
            lead.created_at.strftime("%Y-%m-%d") if lead.created_at else "",
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = d_font; cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"pipeline_{utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/", response_model=list[LeadRead])
def list_leads(
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> list[LeadRead]:
    return [LeadRead.model_validate(item) for item in LeadService(db, current_user).list()]


@router.post("/", response_model=LeadRead, status_code=status.HTTP_201_CREATED)
def create_lead(payload: LeadCreate, db: Session = Depends(get_db)) -> LeadRead:
    return LeadRead.model_validate(LeadService(db).create(payload))


@router.put("/{lead_id}", response_model=LeadRead)
def update_lead_info(
    lead_id: int,
    payload: LeadInfoUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> LeadRead:
    return LeadRead.model_validate(LeadService(db, current_user).update_info(lead_id, payload))


@router.delete("/{lead_id}")
def delete_lead(
    lead_id: int,
    current_user: CurrentUser = Depends(require_roles(RoleCode.SUPER_ADMIN.value)),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    LeadService(db, current_user).delete(lead_id)
    return {"message": "Lead deleted"}


@router.patch("/{lead_id}", response_model=LeadRead)
def update_lead_status(
    lead_id: int,
    payload: LeadUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> LeadRead:
    return LeadRead.model_validate(LeadService(db, current_user).update_status(lead_id, payload))
