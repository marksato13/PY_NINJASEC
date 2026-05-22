import io

import openpyxl
from app.core.time_utils import utcnow
from fastapi import APIRouter, Depends, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.common.deps.auth import CurrentUser, require_roles
from app.common.deps.database import get_db
from app.db.models.enums import RoleCode
from app.modules.clients.schemas import ClientAccessRead, ClientCreate, ClientRead, ClientUpdate
from app.modules.clients.services.client_service import ClientService

router = APIRouter(prefix="/clients", tags=["clients"])


@router.get("/export-xlsx")
def export_clients_xlsx(
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    clients = ClientService(db, current_user).list()

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1E293B")
    h_font = Font(bold=True, color="F8FAFC", name="Calibri", size=11)
    d_font = Font(name="Calibri", size=10)
    a_fill = PatternFill("solid", fgColor="F1F5F9")

    columns = [
        ("ID", 6), ("Empresa", 28), ("Sector", 18), ("Tamaño", 12),
        ("Ciudad", 16), ("País", 14), ("Estado comercial", 16),
        ("Dispositivos", 12), ("Tickets abiertos", 14), ("Notas", 36),
    ]
    for i, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = h_font; cell.fill = h_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_num, client in enumerate(clients, 2):
        fill = a_fill if row_num % 2 == 0 else None
        status_value = client.commercial_status.value if hasattr(client.commercial_status, "value") else str(client.commercial_status)
        values = [
            client.id,
            client.company_name,
            client.sector or "",
            client.size or "",
            client.city or "",
            client.country or "",
            status_value,
            getattr(client, "devices_count", 0),
            getattr(client, "open_tickets_count", 0),
            client.notes or "",
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = d_font; cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"clientes_{utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/", response_model=list[ClientRead])
def list_clients(
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> list[ClientRead]:
    return [ClientRead.model_validate(item) for item in ClientService(db, current_user).list()]


@router.post("/", response_model=ClientRead, status_code=status.HTTP_201_CREATED)
def create_client(
    payload: ClientCreate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> ClientRead:
    return ClientRead.model_validate(ClientService(db, current_user).create(payload))


@router.get("/{client_id}", response_model=ClientRead)
def get_client(
    client_id: int,
    current_user: CurrentUser = Depends(
        require_roles(
            RoleCode.SUPER_ADMIN.value,
            RoleCode.ADMIN.value,
            RoleCode.COLLABORATOR.value,
            RoleCode.CLIENT.value,
        )
    ),
    db: Session = Depends(get_db),
) -> ClientRead:
    return ClientRead.model_validate(ClientService(db, current_user).get_or_404(client_id))


@router.put("/{client_id}", response_model=ClientRead)
def update_client(
    client_id: int,
    payload: ClientUpdate,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> ClientRead:
    return ClientRead.model_validate(ClientService(db, current_user).update(client_id, payload))


@router.get("/{client_id}/access", response_model=list[ClientAccessRead])
def list_client_access(
    client_id: int,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> list[ClientAccessRead]:
    return ClientService(db, current_user).list_access(client_id)


@router.delete("/{client_id}")
def delete_client(
    client_id: int,
    current_user: CurrentUser = Depends(
        require_roles(RoleCode.SUPER_ADMIN.value, RoleCode.ADMIN.value)
    ),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    ClientService(db, current_user).delete(client_id)
    return {"message": "Client deleted"}
